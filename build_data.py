#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_data.py — regenerate data.js for the Stadium Race animation from the
campaign spreadsheet (Fat_Muscle_Measurements.xlsx).

WHY THIS EXISTS
---------------
The race animation reads window.RACE from data.js. data.js is a BUILD ARTIFACT:
do not hand-edit it. Instead you update the spreadsheet each measurement round,
re-run this script, and push. The script:
  * reads the Baseline sheet  -> each athlete's baseline body-fat % / muscle %
  * reads every "<n>th Progress (...)" sheet -> that round's measurements
  * auto-detects how many rounds have data (LATEST) so the race re-ranks and
    extends itself with zero code changes
  * writes data.js (athletes, scores and per-round ranks are computed in the
    browser from this data, exactly as before)

Things NOT in the spreadsheet live in CONFIG below (jersey colours, hair style,
sex, and the planned date schedule). Edit CONFIG only if those change.

USAGE
-----
    python build_data.py                       # reads data/Fat_Muscle_Measurements.xlsx
    python build_data.py --xlsx "C:\\path\\to\\Fat_Muscle_Measurements.xlsx"
    python build_data.py --out data.js         # change output path

After running:  git add -A && git commit -m "Round N data" && git push
(Vercel redeploys automatically; the landing-page iframe and /race.html both
update from the same data.js.)
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required.  Install it with:  pip install openpyxl")

# ----------------------------------------------------------------------------
# CONFIG — presentation + schedule data that is NOT stored in the spreadsheet.
# ----------------------------------------------------------------------------

# Women in the campaign; everyone else defaults to 'm'.
FEMALE = {"Orawan", "Maneenund", "Piyathida"}

# Per-athlete jersey colour + hair-style index (matches the campaign website).
STYLE = {
    "Pipat":     {"color": "#06B6D4", "hair": 0},
    "Best":      {"color": "#14B8A6", "hair": 1},
    "Ben":       {"color": "#EC4899", "hair": 2},
    "Orawan":    {"color": "#F97316", "hair": 0},
    "Nott":      {"color": "#EF4444", "hair": 3},
    "Maneenund": {"color": "#10B981", "hair": 1},
    "Bob":       {"color": "#F59E0B", "hair": 4},
    "Ann":       {"color": "#6366F1", "hair": 5},
    "Ohm":       {"color": "#3B82F6", "hair": 6},
    "Sonny":     {"color": "#22D3EE", "hair": 7},
    "Wason":     {"color": "#FB923C", "hair": 8},
    "Piyathida": {"color": "#84CC16", "hair": 2},
    "Pruang":    {"color": "#A855F7", "hair": 9},
    "Jue":       {"color": "#8B5CF6", "hair": 10},
}

# Total rounds the track is drawn for (P1..P8). The race shows future rounds as
# empty until their sheet appears.
NUM_STAGES = 8

# Finish-line goal the leader is chasing (combined score).
FINISH_SCORE = 6.0

# Planned schedule. Index 0 = Baseline, 1..8 = progress rounds.
# If an actual measurement date differs from the plan, edit the matching entry.
STAGE_LABELS = ["Baseline", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]
STAGE_DATES = ["16 Mar", "31 Mar", "20 Apr", "5 May", "19 May",
               "8 Jun", "22 Jun", "6 Jul", "20 Jul"]
STAGE_FULL = [
    "Baseline · 16 Mar 2026",
    "1st progress · 31 Mar 2026",
    "2nd progress · 20 Apr 2026",
    "3rd progress · 5 May 2026",
    "4th progress · 19 May 2026",
    "5th progress · 8 Jun 2026",
    "6th progress · 22 Jun 2026",
    "7th progress · 6 Jul 2026",
    "Final · 20 Jul 2026",
]

# ----------------------------------------------------------------------------
# Frozen model logic. This is appended verbatim to the generated data.js; it
# turns the data above into window.RACE. LATEST is derived from the rounds that
# actually have data, so the race re-ranks itself each round automatically.
# ----------------------------------------------------------------------------
MODEL_JS = """
  // ----- model (frozen; computed in the browser) -----------------------------
  const NAMES = Object.keys(BASE);
  // Most recent round that has data — drives ranking and the "now" marker.
  const LATEST = Object.keys(RAW).length
    ? Math.max(...Object.keys(RAW).map(Number)) : 0;
  const score = (b, fat, mus) => (b.fat - fat) + (mus - b.muscle);

  // Per-stage record for each athlete (stage 0 = Baseline, 1..NUM_STAGES = rounds).
  // Missed rounds carry forward the athlete's last-known reading; rounds with no
  // sheet yet are marked future (null) and drawn as empty track.
  const athletes = NAMES.map((name) => {
    const b = BASE[name];
    const stages = [];
    let last = { fat: b.fat, muscle: b.muscle };
    stages.push({ fat: b.fat, muscle: b.muscle, carried: false, present: true, score: 0 });
    for (let s = 1; s <= NUM_STAGES; s++) {
      if (s <= LATEST) {
        const m = RAW[s] && RAW[s][name];
        if (m) {
          last = { fat: m[0], muscle: m[1] };
          stages.push({ fat: m[0], muscle: m[1], carried: false, present: true,
                        score: score(b, m[0], m[1]) });
        } else {
          stages.push({ fat: last.fat, muscle: last.muscle, carried: true, present: false,
                        score: score(b, last.fat, last.muscle) });
        }
      } else {
        stages.push({ fat: null, muscle: null, carried: false, present: false, score: null, future: true });
      }
    }
    return { name, sex: b.sex, base: b, style: STYLE[name], stages };
  });

  // Ranks per measured stage (1 = best). Ties broken by fat lost, then name.
  const ranks = [];
  for (let s = 0; s <= LATEST; s++) {
    const order = athletes
      .map((a) => ({ name: a.name, sc: a.stages[s].score,
                     fatLost: a.base.fat - a.stages[s].fat }))
      .sort((x, y) => (y.sc - x.sc) || (y.fatLost - x.fatLost) || x.name.localeCompare(y.name));
    const r = {};
    order.forEach((o, i) => { r[o.name] = i + 1; });
    ranks.push(r);
  }

  window.RACE = {
    athletes, ranks, NUM_STAGES, LATEST, FINISH_SCORE, NAMES,
    STAGE_LABELS, STAGE_DATES, STAGE_FULL,
  };
})();
"""

# Rows that are headers/labels, not athlete data.
NON_NAME = {"name", "group average", ""}
ORDINAL_RE = re.compile(r"^\s*(\d+)\s*(st|nd|rd|th)\b", re.IGNORECASE)


def jsnum(v):
    """Render a measurement as a compact JS number (74.0 -> 74, 22.75 -> 22.75)."""
    f = float(v)
    return str(int(f)) if f == int(f) else repr(round(f, 4))


def is_name_row(cell):
    return isinstance(cell, str) and cell.strip().lower() not in NON_NAME \
        and "campaign" not in cell.lower()


def read_rows(ws):
    """Yield (name, [floats...]) for every athlete data row in a sheet."""
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        name = row[0]
        if not is_name_row(name):
            continue
        yield name.strip(), list(row[1:])


def parse_baseline(ws):
    """Baseline sheet -> {name: (avg_fat, avg_muscle)} using the Avg columns
    (F,G) when present, else the first measurement (B,C)."""
    base = {}
    for name, cols in read_rows(ws):
        # cols index (0-based, relative to col B): Avg BF% = 4, Avg Muscle% = 5
        fat = cols[4] if len(cols) > 5 and cols[4] not in (None, "") else cols[0]
        mus = cols[5] if len(cols) > 5 and cols[5] not in (None, "") else cols[1]
        if fat in (None, "") or mus in (None, ""):
            continue
        base[name] = (float(fat), float(mus))
    return base


def parse_progress(ws):
    """Progress sheet -> {name: (fat, muscle)} skipping blank/absent rows."""
    out = {}
    for name, cols in read_rows(ws):
        fat = cols[0] if len(cols) > 0 else None
        mus = cols[1] if len(cols) > 1 else None
        if fat in (None, "") or mus in (None, ""):
            continue  # present in sheet but no reading = treated as missed
        out[name] = (float(fat), float(mus))
    return out


def build(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Locate the baseline sheet and the numbered progress sheets.
    baseline_sheet = None
    progress = {}  # stage_number -> sheet name
    for sn in wb.sheetnames:
        low = sn.strip().lower()
        if low.startswith("baseline"):
            baseline_sheet = sn
        else:
            m = ORDINAL_RE.match(sn.strip())
            if m and "progress" in low:
                progress[int(m.group(1))] = sn

    if baseline_sheet is None:
        sys.exit("No 'Baseline...' sheet found. Sheets: %s" % wb.sheetnames)

    base_raw = parse_baseline(wb[baseline_sheet])
    names = list(base_raw.keys())  # roster + display order from the baseline sheet

    # Build BASE with sex + baseline values.
    base = {n: {"sex": "f" if n in FEMALE else "m",
                "fat": base_raw[n][0], "muscle": base_raw[n][1]} for n in names}

    # Build RAW for every round that has a sheet, validating names.
    raw = {}
    for stage in sorted(progress):
        if stage > NUM_STAGES:
            print("  ! skipping round %d — beyond NUM_STAGES=%d (extend CONFIG)"
                  % (stage, NUM_STAGES))
            continue
        rows = parse_progress(wb[progress[stage]])
        unknown = [n for n in rows if n not in base]
        for n in unknown:
            print("  ! round %d: unknown athlete '%s' (not in Baseline) — skipped" % (stage, n))
            rows.pop(n)
        raw[stage] = rows

    # Sanity warnings.
    missing_style = [n for n in names if n not in STYLE]
    if missing_style:
        print("  ! no STYLE (colour/hair) for: %s — add to CONFIG" % ", ".join(missing_style))

    write_js(base, raw, out_path, baseline_sheet, progress, names)
    return base, raw


def write_js(base, raw, out_path, baseline_sheet, progress, names):
    L = []
    L.append("/* ============================================================")
    L.append("   Fat & Muscle Campaign — Stadium Race data model")
    L.append("   GENERATED FILE — do not edit by hand.")
    L.append("   Edit the spreadsheet, then run:  python build_data.py")
    L.append("   Combined score = (baselineFat - fat) + (muscle - baselineMuscle)")
    L.append("   Missed rounds carry forward the last-known reading.")
    L.append("   ============================================================ */")
    L.append("(function () {")
    L.append("  // Baseline = avg body-fat %% / muscle %% from sheet '%s'." % baseline_sheet)
    L.append("  const BASE = {")
    for n in names:
        b = base[n]
        L.append("    %s: { sex: '%s', fat: %s, muscle: %s }," %
                 (jskey(n), b["sex"], jsnum(b["fat"]), jsnum(b["muscle"])))
    L.append("  };")
    L.append("")
    L.append("  // Raw measurements per round (omitted name = missed that round).")
    L.append("  const RAW = {")
    for stage in sorted(raw):
        present = raw[stage]
        absent = [n for n in names if n not in present]
        note = ("  // all present" if not absent
                else "  // missed: " + ", ".join(absent))
        L.append("    %d: {%s" % (stage, note))
        # emit in roster order for stable diffs
        line = "      "
        for n in names:
            if n in present:
                f, mus = present[n]
                tok = "%s:[%s,%s], " % (jskey(n), jsnum(f), jsnum(mus))
                if len(line) + len(tok) > 92:
                    L.append(line.rstrip())
                    line = "      "
                line += tok
        if line.strip():
            L.append(line.rstrip())
        L.append("    },")
    L.append("  };")
    L.append("")
    L.append("  const STYLE = {")
    for n in names:
        s = STYLE.get(n, {"color": "#888888", "hair": 0})
        L.append("    %s: { color: '%s', hair: %d }," % (jskey(n), s["color"], s["hair"]))
    L.append("  };")
    L.append("")
    L.append("  const NUM_STAGES = %d;" % NUM_STAGES)
    L.append("  const FINISH_SCORE = %s;" % repr(FINISH_SCORE))
    L.append("  const STAGE_LABELS = %s;" % js_arr(STAGE_LABELS))
    L.append("  const STAGE_DATES = %s;" % js_arr(STAGE_DATES))
    L.append("  const STAGE_FULL = [")
    for s in STAGE_FULL:
        L.append("    %s," % js_str(s))
    L.append("  ];")
    L.append(MODEL_JS.rstrip("\n"))
    L.append("")

    Path(out_path).write_text("\n".join(L), encoding="utf-8")


def jskey(name):
    """Object key — bare identifier if safe, else quoted."""
    return name if re.match(r"^[A-Za-z_$][A-Za-z0-9_$]*$", name) else js_str(name)


def js_str(s):
    return "'" + str(s).replace("\\", "\\\\").replace("'", "\\'") + "'"


def js_arr(items):
    return "[" + ", ".join(js_str(i) for i in items) + "]"


def main():
    here = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(description="Regenerate data.js from the campaign spreadsheet.")
    ap.add_argument("--xlsx", default=str(here / "data" / "Fat_Muscle_Measurements.xlsx"),
                    help="path to Fat_Muscle_Measurements.xlsx (default: data/ in repo)")
    ap.add_argument("--out", default=str(here / "data.js"),
                    help="output path (default: data.js in repo)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit("Spreadsheet not found: %s\nPass --xlsx <path> or copy the file to data/." % xlsx)

    print("Reading: %s" % xlsx)
    base, raw = build(xlsx, args.out)
    rounds = sorted(raw)
    print("  athletes: %d" % len(base))
    print("  rounds with data: %s  (LATEST = %d)" %
          (", ".join("P%d(%d)" % (s, len(raw[s])) for s in rounds) or "none",
           max(rounds) if rounds else 0))
    print("Wrote: %s" % args.out)
    print("Next:  git add -A && git commit -m \"Round data update\" && git push")


if __name__ == "__main__":
    main()
