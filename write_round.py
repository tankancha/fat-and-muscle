#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
write_round.py — add or replace ONE progress round in the campaign spreadsheet.

This is the spreadsheet-writing half of the measurement pipeline. The
/update-round Claude command does the OCR + your review, then calls this with a
small JSON payload of confirmed readings. It writes a "<n>th Progress (<date>)"
sheet, which build_data.py then turns into data.js (progress, leaderboard, race).

PAYLOAD (JSON file passed via --json):
    {
      "round": 5,
      "date_label": "8 Jun 26",          // goes in the sheet name / tab
      "title_date": "8 Jun 2026",        // goes in the sheet's title row
      "readings": { "Wason": [29.0, 67.2], "Pruang": [19.0, 76.8] }
                   //  roster name : [body_fat_pct, muscle_pct]
    }

USAGE:
    python write_round.py --json data/_round.json
    python write_round.py --json data/_round.json --xlsx "C:\\path\\to\\file.xlsx"

Names are validated against the roster (the Baseline sheet). Unknown names abort.
The target spreadsheet must be CLOSED in Excel or the save will fail.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

DEFAULT_XLSX = r"C:\Users\Admin\OneDrive\Claude Cowork\03_Projects\Fat and Muscle\Fat and Muscle\Fat_Muscle_Measurements.xlsx"

NON_NAME = {"name", "group average", ""}


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def roster_from_baseline(wb):
    """Read the athlete roster (names) from the Baseline sheet."""
    sheet = next((s for s in wb.sheetnames if s.strip().lower().startswith("baseline")), None)
    if not sheet:
        sys.exit("No 'Baseline...' sheet found in the workbook.")
    names = []
    for row in wb[sheet].iter_rows(values_only=True):
        a = row[0]
        if isinstance(a, str) and a.strip().lower() not in NON_NAME and "campaign" not in a.lower():
            names.append(a.strip())
    return names


def main():
    ap = argparse.ArgumentParser(description="Write one progress round into the spreadsheet.")
    ap.add_argument("--json", required=True, help="path to the round payload JSON")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="spreadsheet to update (default: OneDrive source)")
    args = ap.parse_args()

    payload = json.loads(Path(args.json).read_text(encoding="utf-8"))
    rnd = int(payload["round"])
    date_label = str(payload["date_label"]).strip()
    title_date = str(payload.get("title_date", date_label)).strip()
    readings = payload["readings"]
    if not readings:
        sys.exit("Payload has no readings.")

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit("Spreadsheet not found: %s" % xlsx)

    try:
        wb = openpyxl.load_workbook(xlsx)
    except PermissionError:
        sys.exit("Cannot open %s — close it in Excel and try again." % xlsx)

    roster = roster_from_baseline(wb)
    roster_lower = {n.lower(): n for n in roster}

    # Validate + normalise names to exact roster spelling.
    clean = {}
    unknown = []
    for name, vals in readings.items():
        canon = roster_lower.get(str(name).strip().lower())
        if not canon:
            unknown.append(name)
            continue
        bf, mu = float(vals[0]), float(vals[1])
        clean[canon] = (bf, mu)
    if unknown:
        sys.exit("These names are not in the roster (fix name_map or the payload): %s\nRoster: %s"
                 % (", ".join(map(str, unknown)), ", ".join(roster)))

    sheet_name = "%s Progress (%s)" % (ordinal(rnd), date_label)

    # Replace an existing sheet for this round (match by leading ordinal), so
    # re-running the round is idempotent.
    for s in list(wb.sheetnames):
        st = s.strip().lower()
        if st.startswith(ordinal(rnd).lower()) and "progress" in st:
            del wb[s]

    # Insert just before a "Comparison" sheet if present, else append.
    idx = wb.sheetnames.index("Comparison") if "Comparison" in wb.sheetnames else None
    ws = wb.create_sheet(sheet_name) if idx is None else wb.create_sheet(sheet_name, idx)

    ws.append(["Fat & Muscle Campaign — %s Progress Measurements (%s)" % (ordinal(rnd), title_date)])
    ws.append(["Name", "Body Fat %", "Muscle %"])
    for n in roster:                       # roster order, only those measured
        if n in clean:
            bf, mu = clean[n]
            ws.append([n, bf, mu])
    ws.append(["GROUP AVERAGE", 0, 0])

    try:
        wb.save(xlsx)
    except PermissionError:
        sys.exit("Cannot save %s — close it in Excel and try again." % xlsx)

    absent = [n for n in roster if n not in clean]
    print("Wrote sheet '%s' to %s" % (sheet_name, xlsx.name))
    print("  measured (%d): %s" % (len(clean), ", ".join(n for n in roster if n in clean)))
    print("  absent  (%d): %s" % (len(absent), ", ".join(absent) or "none"))


if __name__ == "__main__":
    main()
